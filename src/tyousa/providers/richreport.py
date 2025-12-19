from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from tyousa.models import StatsResult, StatsSnapshot

logger = logging.getLogger(__name__)


ROW_LABELS = {
    "一般世帯総数": "households_total",
    "単身世帯数": "one_person",
    "２人世帯数": "two_person",
    "３人世帯数": "three_person",
    "４人世帯数": "four_person",
    "５人世帯数": "five_person",
    "６人以上世帯数": "six_plus",
    "主世帯数": "main_households",
    "共同住宅世帯数": "apartment_households",
    "住宅に住む一般世帯": "housing_households",
    "持ち家世帯数": "owner_households",
    "民営の借家世帯数": "private_rental_households",
}


class RichReportProvider:
    def __init__(self, path: Path) -> None:
        self.path = path

    def fetch(self) -> StatsResult:
        book = load_workbook(self.path, data_only=True)
        if "世帯数" not in book.sheetnames:
            raise ValueError("RichReport is missing '世帯数' sheet")
        sheet = book["世帯数"]

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        area_columns: dict[int, int] = {}
        for idx, value in enumerate(header_row):
            if value and "１次" in str(value):
                area_columns[600] = idx
            if value and "２次" in str(value):
                area_columns[2000] = idx
        if not area_columns:
            raise ValueError("RichReport missing １次/２次 area columns")

        metrics_by_radius: dict[int, dict[str, int]] = {600: {}, 2000: {}}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            label = str(row[0]).strip() if row[0] else None
            if not label or label not in ROW_LABELS:
                continue
            for radius, col_idx in area_columns.items():
                value = row[col_idx]
                if value is None:
                    continue
                try:
                    metrics_by_radius[radius][ROW_LABELS[label]] = int(value)
                except (TypeError, ValueError):
                    logger.warning("Non-numeric value for %s (%s): %s", label, radius, value)

        snapshots: dict[int, StatsSnapshot] = {}
        for radius, values in metrics_by_radius.items():
            missing = [v for v in ROW_LABELS.values() if v not in values]
            if missing:
                raise ValueError(f"RichReport missing values for {missing} in {radius}m column")
            snapshots[radius] = StatsSnapshot(**values)

        return StatsResult(by_radius=snapshots)
