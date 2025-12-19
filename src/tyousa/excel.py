from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from tyousa.models import CandidateMetrics

logger = logging.getLogger(__name__)


COLUMN_MAPPING = {
    "编号": "id",
    "区域名称": "address",
    "纬度": "lat",
    "经度": "lon",
    "600m内户数": "households_600",
    "2km内户数": "households_2000",
    "600m内租赁户占比": "rental_share_600",
    "2km内租赁户占比": "rental_share_2000",
    "600m内集合住宅占比": "apartment_share_600",
    "600m内小户型占比（单身+2人户）": "small_household_share_600",
    "2km内家庭户占比": "family_share_2000",
    "人口趋势指数": "trend_index",
    "600m内竞品数量": "count_competitors_600",
    "2km内竞品数量": "count_competitors_2000",
    "最近竞品距离(米)": "nearest_competitor_distance_m",
    "到最近车站距离(米)": "nearest_station_distance_m",
    "到生活锚点距离(米：超市/药妆/家居)": "nearest_anchor_distance_m",
    "到主干道距离(米)": "main_road_distance_m",
    "300m内停车锚点(0/1)": "parking_anchor_300m",
    "600m内强竞品(0/1)": "strong_competitor_600",
    "2km内强竞品(0/1)": "strong_competitor_2000",
}


class ExcelWriter:
    def __init__(self, template_path: Path, preserve_manual: bool = True) -> None:
        self.template_path = template_path
        self.preserve_manual = preserve_manual

    def write(self, metrics: list[CandidateMetrics], output_path: Path) -> None:
        book = load_workbook(self.template_path)
        if "候选点" not in book.sheetnames:
            raise ValueError("Template missing '候选点' sheet")
        sheet = book["候选点"]

        header_map = self._header_map(sheet)
        start_row = self._find_start_row(sheet)

        for offset, item in enumerate(metrics):
            row_idx = start_row + offset
            self._write_row(sheet, row_idx, item, header_map)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        book.save(output_path)

    def _header_map(self, sheet: Worksheet) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(sheet[1], start=1):
            header = str(cell.value).strip() if cell.value is not None else None
            if header:
                header_map[header] = idx
        return header_map

    def _find_start_row(self, sheet: Worksheet, base_row: int = 5) -> int:
        row = base_row
        while True:
            cell = sheet.cell(row=row, column=1)
            if cell.value in (None, ""):
                return row
            row += 1

    def _write_row(
        self, sheet: Worksheet, row_idx: int, metrics: CandidateMetrics, header_map: dict[str, int]
    ) -> None:
        values = self._flatten_metrics(metrics)
        for header, attr in COLUMN_MAPPING.items():
            if header not in header_map:
                continue
            col_idx = header_map[header]
            value = values.get(attr)
            cell = sheet.cell(row=row_idx, column=col_idx)
            if self.preserve_manual and cell.value not in (None, ""):
                continue
            if value is None:
                continue
            cell.value = value

    def _flatten_metrics(self, metrics: CandidateMetrics) -> dict[str, float | None]:
        return {
            "id": metrics.candidate_id,
            "address": metrics.address,
            "lat": metrics.lat,
            "lon": metrics.lon,
            "households_600": metrics.stats.households_600,
            "households_2000": metrics.stats.households_2000,
            "rental_share_600": metrics.stats.rental_share_600,
            "rental_share_2000": metrics.stats.rental_share_2000,
            "apartment_share_600": metrics.stats.apartment_share_600,
            "small_household_share_600": metrics.stats.small_household_share_600,
            "family_share_2000": metrics.stats.family_share_2000,
            "trend_index": metrics.stats.trend_index,
            "count_competitors_600": metrics.poi.count_competitors_600,
            "count_competitors_2000": metrics.poi.count_competitors_2000,
            "nearest_competitor_distance_m": metrics.poi.nearest_competitor_distance_m,
            "nearest_station_distance_m": metrics.poi.nearest_station_distance_m,
            "nearest_anchor_distance_m": metrics.poi.nearest_anchor_distance_m,
            "main_road_distance_m": metrics.poi.main_road_distance_m,
            "parking_anchor_300m": metrics.poi.parking_anchor_300m,
            "strong_competitor_600": metrics.poi.strong_competitor_600,
            "strong_competitor_2000": metrics.poi.strong_competitor_2000,
        }
